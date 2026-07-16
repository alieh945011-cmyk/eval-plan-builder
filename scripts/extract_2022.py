# 2022개정 전과목 xlsx에서 성취기준·성취수준을 추출해 public/data/2022/{과목}.json 생성
import json
import re
import sys
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\user\Documents\1. 클로드코드\26-07-16. 수행평가 제작 프로그램\2022개정 교육과정 성취수준(중)\pdf\(전과목)2022 교육과정_중학교_교과별_성취기준 및 성취수준(전과목)-역사는 없음.xlsx")
OUT_DIR = Path(__file__).resolve().parent.parent / "public" / "data" / "2022"

CODE_RE = re.compile(r"^\s*\[([0-9]{1,2}[가-힣]{1,4}[0-9]{2}-[0-9]{2})\]\s*(.*)$", re.S)


def norm(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None


def sheet_pairs(wb):
    """시트 이름 표기 편차(공백·언더스코어)를 흡수해 과목별 (성취기준별, 영역별) 쌍을 만든다."""
    std, dom = {}, {}
    for name in wb.sheetnames:
        key = name.replace(" ", "").replace("_", "")
        if key.endswith("성취기준별"):
            std[key[: -len("성취기준별")]] = name
        elif key.endswith("영역별"):
            dom[key[: -len("영역별")]] = name
    return {subj: (std[subj], dom.get(subj)) for subj in std}


DOMAIN_RE = re.compile(r"^\(\d+\)\s*")  # "(1) 수와 연산" 형태의 영역 표기


def extract_standards(ws):
    """행 배치가 과목마다 달라 규칙으로 흡수한다.
    - 반복 헤더 행('영역'/'성취기준'/'성취기준별 성취수준')은 건너뜀
    - 1열 값이 "(n) …" 패턴이면 영역, 아니면 단원(성취기준 행에 붙는 소분류)
    - 같은 이름의 영역이 연속되면 병합
    """
    domains = []  # [{name, standards:[{code,text,unit,levels:{}}]}]
    cur_domain = None
    cur_unit = None
    cur_std = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        domain, standard, level, desc = (norm(c) for c in row[:4])
        if standard == "성취기준" or domain == "영역" or level == "성취기준별 성취수준":
            continue  # 반복 헤더
        if domain:
            if DOMAIN_RE.match(domain) or (not standard and not level):
                name = domain
                if cur_domain is None or cur_domain["name"] != name:
                    cur_domain = {"name": name, "standards": []}
                    domains.append(cur_domain)
                cur_unit = None
                if not standard and not level:
                    continue
            else:
                cur_unit = domain  # 단원 표기 (예: "1. 소인수분해")
        if standard:
            m = CODE_RE.match(standard)
            if m:
                code, text = m.group(1), m.group(2).strip()
            else:
                code, text = None, standard
            cur_std = {"code": code, "text": text, "unit": cur_unit, "levels": {}}
            if cur_domain is None:
                cur_domain = {"name": "", "standards": []}
                domains.append(cur_domain)
            cur_domain["standards"].append(cur_std)
        if level and desc and cur_std is not None:
            cur_std["levels"][level] = desc
    # 영역 이름이 반복 기입된 과목(국어 등)은 연속 병합 후에도 중복될 수 있어 재병합
    merged = []
    for d in domains:
        if merged and merged[-1]["name"] == d["name"]:
            merged[-1]["standards"].extend(d["standards"])
        else:
            merged.append(d)
    return merged


def fill_merged_levels(domains, order="ABCDE"):
    """원자료(보급본 PDF)는 인접 수준을 병합 셀로 묶어 하나의 진술을 공유한다.
    xlsx에는 병합 하단 셀이 빈칸으로 옮겨져 있으므로, 빈 수준은 바로 위 수준의
    진술을 그대로 물려받고 mergedUp 표시를 남긴다(원자료 병합 구조 보존)."""
    for d in domains:
        for s in d["standards"]:
            present = [lv for lv in order if lv in s["levels"]]
            if not present:
                continue
            merged_up = []
            prev = None
            for lv in order:
                if lv in s["levels"]:
                    prev = lv
                elif prev is not None:
                    s["levels"][lv] = s["levels"][prev]
                    merged_up.append(lv)
            if merged_up:
                s["mergedUp"] = merged_up


def extract_domain_levels(ws):
    result = {}  # {영역: {A: {범주: 진술}}}
    cur_domain = None
    cur_level = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        domain, level, category, desc = (norm(c) for c in row[:4])
        if domain:
            cur_domain = domain
            result.setdefault(cur_domain, {})
        if level:
            cur_level = level
        if category and desc and cur_domain and cur_level:
            result[cur_domain].setdefault(cur_level, {})[category] = desc
    return result


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    index = []
    for subj, (std_sheet, dom_sheet) in sheet_pairs(wb).items():
        domains = extract_standards(wb[std_sheet])
        dom_levels = extract_domain_levels(wb[dom_sheet]) if dom_sheet else {}
        # 영역별 성취수준을 영역명 유사 매칭으로 붙임 (중점(･)·공백 표기 편차 흡수)
        def dkey(s):
            return re.sub(r"[\s･·.()0-9]", "", s or "")
        dom_levels_by_key = {dkey(k): v for k, v in dom_levels.items()}
        for d in domains:
            d["domainLevels"] = dom_levels_by_key.get(dkey(d["name"]), {})
        levels = sorted({lv for d in domains for s in d["standards"] for lv in s["levels"]})
        levels = [lv for lv in levels if lv in "ABCDE"]
        fill_merged_levels(domains, order="".join(levels))
        data = {
            "curriculum": "2022",
            "subject": subj,
            "levels": levels,
            "domains": domains,
        }
        out = OUT_DIR / f"{subj}.json"
        out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
        n_std = sum(len(d["standards"]) for d in domains)
        n_missing = sum(1 for d in domains for s in d["standards"] if len(s["levels"]) < len(levels))
        index.append({"subject": subj, "standards": n_std, "levels": "".join(levels),
                      "domains": len(domains), "incompleteLevelSets": n_missing})
        print(f"{subj}: 영역 {len(domains)}, 성취기준 {n_std}, 수준 {''.join(levels)}, 수준누락 {n_missing}")
    (OUT_DIR / "_index.json").write_text(json.dumps(index, ensure_ascii=False, indent=1), encoding="utf-8")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
